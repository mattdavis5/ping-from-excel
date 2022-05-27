#This program will ping each row of an Excel spreadsheet that contains a value

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import colors, Font, Color
from pythonping import ping

while True:
    #Input the file path of the Excel file
    file_location = input("Enter the Excel file's path: \n")
    print("\nConfirmed path is " + file_location)

    #Load the excel workbook
    try:
        excel_workbook = openpyxl.load_workbook(file_location)
    except InvalidFileException as ife:
        print("\nPlease enter a valid file path. \nAvailable formats are .xlsx, .xlsm, .xltx, .xltm\n")
    except FileNotFoundError as fnf:
        print("\nPlease enter an existing file path.\n")
    else:
        break

#Sets the first/only Excel sheet in the workbook as the active sheet for further processing in this program
excel_sheet = excel_workbook.active

#Input the column letter to ping
column = input("\nEnter the column to ping: ")
print("\nColumn " + column + " will be tested")

#Convert the column letter to a column index (example: A = 1, B = 2)
column = openpyxl.utils.column_index_from_string(column)

#Input which row of Excel sheet to begin testing cells
while True:
    try:
        min_row = int(input("\nEnter the first row with pingable values: "))
        print("\nConfirmed starting row is " + str(min_row))
    except:
        print("\nPlease enter an integer for the starting row")
    else:
        break


#Find the last row which contains cell values
max_row = excel_sheet.max_row

#Find the number of hosts that will be pinged
host_count = max_row - min_row + 1
print("\nThe number of hosts to run the ping test are: " + str(host_count))

#Input file location of log file
while True:
    log_location = input("\nEnter the log .txt file's path: \n")
    print("\nConfirmed log file path is " + log_location)

    #Write to log file, create one if does not already exist
    try:
        with open(log_location, 'w') as f:
            f.write("Log file for ping test from Excel file - " + file_location)
    except:
        print("\nPlease enter a valid path for the .txt log file")
    else:
        break

#Append ping test statement to log file
with open(log_location, 'a') as f:
    f.write("\n\nPinging " + str(host_count) + " hosts\n")

#Counter variable of pingable hosts
ping_success_count = 0

#Ratio variable of pingable hosts to all hosts
ping_success_ratio = 0

print("\nRunning the ping test...")

#For each row in the Excel sheet, and each cell in the row, ping the host
for rows in excel_sheet.iter_rows(min_row=min_row, max_col=column, max_row=max_row):
     for cell in rows:
         host = cell.value

        #ping_result is an object created from the ping() method
         ping_result = ping(host)

        #Append ping result to log file
         with open(log_location, 'a') as f:
            f.write("\n\n\n\nPinging " + host + " at " + str(cell))
            f.write("\n\n" + str(ping_result) + "\n")
            f.write("\n-------------------------------------------------------------")
         
         #If the ping result is successful, change cell font to green and update success counter variable
         if ping_result.success():
            cell.font = Font(color="0000FF00", italic=False)
            ping_success_count += 1
         #If ping result is not successful, change cell font to red, italicize font
         else:
            cell.font = Font(color="00FF0000", italic=True)


#Save the Excel workbook to the file location entered previously
excel_workbook.save(file_location)

print("\n" + str(ping_success_count) + " hosts are reachable")

#Calculate the percentage of tested hosts that are reachable
ping_success_ratio = round((ping_success_count / host_count)*100,2)
print("\n" + str(ping_success_ratio) + "% of hosts are reachable")

#Append ping summary to log file
with open(log_location, 'a') as f:
    f.write("\n\n\nSummary:\n" + str(ping_success_count) + " hosts are reachable")
    f.write("\n" + str(ping_success_ratio) + "% of hosts are reachable")

print("\nPing test is complete. Please review the log file for detailed ping results, and Excel file for text modifications. \nHosts that are pingable are in standard green text, unreachable hosts are in italic red text in the Excel file.")

